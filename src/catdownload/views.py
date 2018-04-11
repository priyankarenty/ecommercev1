import csv
import xlwt
from django.http import HttpResponse
from products.models import Product
from ProductAccessory.models import ProductAccessory
from ProductAVEng.models import ProductAVEng
import openpyxl
from openpyxl.cell import get_column_letter
from django.db.models.query import QuerySet
from datetime import datetime
current_month = datetime.now().month


def export_hardware_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Hardware_Catalog.csv"'

    writer = csv.writer(response)
    writer.writerow(['MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions'])

    products = Product.objects.all().filter(DateofEntry__month=current_month).values_list('MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions')
    for user in products:
        writer.writerow(user)

    return response

def export_accessory_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Accessory_Catalog.csv"'

    writer = csv.writer(response)
    writer.writerow(['MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions'])

    products = ProductAccessory.objects.all().filter(DateofEntry__month=current_month).values_list('MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions')
    for user in products:
        writer.writerow(user)

    return response


def export_aveng_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="AVEng_Catalog.csv"'

    writer = csv.writer(response)
    writer.writerow(['MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions'])

    products = ProductAVEng.objects.all().filter(DateofEntry__month=current_month).values_list('MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions')
    for user in products:
        writer.writerow(user)

    return response

def export_hardware_xlsx(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Hardware_Catalog.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Product')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions',]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = Product.objects.all().filter(DateofEntry__month=current_month).values_list('MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType', 'RAM', 'HardDiskStorage', 'OS', 'Processor', 'TouchScreen', 'ScreenSize', 'ScreenBorder', 'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


def export_accessory_xlsx(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Accessory_Catalog.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Product')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType','Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions',]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = ProductAccessory.objects.all().filter(DateofEntry__month=current_month).values_list('MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType',  'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


def export_aveng_xlsx(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="AVEng_Catalog.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Product')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType','Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions',]
    

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = ProductAVEng.objects.all().filter(DateofEntry__month=current_month).values_list('MonthofEntry', 'DateofEntry', 'ProductName', 'ProductCategory', 'ProductType',  'Warranty', 'WarrantyPeriod', 'Image', 'Supplier', 'StockAvailable', 'SupplierType', 'Partnumber', 'UPC', 'SKU', 'Region', 'Country', 'QuoteValidity', 'UnitofMeasure', 'Currency', 'TotalPrice', 'BasePriceperUnit', 'MarkupPriceperUnit', 'DeliveryChargeperunit', 'WarrantyPrice', 'AssetTagPrice', 'Taxes', 'RecycleFee', 'FreightCharge', 'AnyOtherFee', 'StockLeadTime', 'DeliveryLeadTime', 'PriceChange', 'GoogleApproverNameAvailable', 'GoogleApprover', 'ReasonForPriceChange', 'EoLStatus', 'EoLDate', 'PreviousMPN', 'PreviousPartNumber', 'SplInstructions')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response
